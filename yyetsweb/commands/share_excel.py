#!/usr/local/bin/python3
# coding: utf-8

# YYeTsBot - share_excel.py
# 12/18/21 19:21
#

__author__ = "Benny <benny.think@gmail.com>"

import pathlib
import sys

import openpyxl

web_path = pathlib.Path(__file__).parent.parent.resolve().as_posix()
sys.path.append(web_path)
from tqdm import tqdm

from common.utils import ts_date
from databases.base import Mongo

wb = openpyxl.open("aliyun.xlsx")

data = {}

for ws in wb.worksheets:
    line = 0
    for line in range(1, ws.max_row + 1):
        name = ws.cell(line, 1).value
        link = ws.cell(line, 2).value
        line += 1
        data[name] = link

template = {
    "username": "Benny",
    "ip": "127.0.0.1",
    "date": "",
    "browser": "cli",
    "content": "",
    "resource_id": 234,
    "type": "parent",
}
col = Mongo().db["comment"]
share_doc = {
    "status": 1.0,
    "info": "OK",
    "data": {
        "info": {
            "id": 234,
            "cnname": "网友分享",
            "enname": "",
            "aliasname": "",
            "channel": "share",
            "channel_cn": "",
            "area": "",
            "show_type": "",
            "expire": "1610401225",
            "views": 0,
        },
        "list": [],
    },
}

Mongo().db["yyets"].update_one({"data.info.id": 234}, {"$set": share_doc}, upsert=True)

for name, link in tqdm(data.items()):
    template["content"] = f"{name}\n{link}"
    template["date"] = ts_date()
    col.insert_one(template.copy())
